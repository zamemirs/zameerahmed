import pandas as pd
import numpy as np
import os
import sys
from io import StringIO
from openpyxl import load_workbook
from openpyxl.styles import PatternFill


def savetolocal(df,filename):

 buffer_file = StringIO()
 df.to_csv(buffer_file, index=False)
 data = buffer_file.getvalue() # Convert to bytes
 outputfile = os.path.join(base_directory,f"{filename}.csv")

 save_file = open(outputfile,'w+')
 save_file.write(data)
 buffer_file.close()
 save_file.close()

def is_float(s):
    try:
        float(s)
        return True
    except ValueError:
        return False

def selectrandom(df):
    start_row, end_row = 1, 1160  # Modify these as per your requirement
    start_col, end_col = 5, 26    # Modify these as per your requirement

    # Select rows and columns within the specified range
    df_range = df.iloc[start_row:end_row, start_col:end_col]
    # Replace blank values with np.nan
    df_range = df_range.replace(r'^\s*$', np.nan, regex=True).replace([r''], np.nan)
    
    # Flatten the DataFrame to a Series and remove NaN values
    series = df_range.stack()
    #print (series)
    df_range_size = len(series)  # Assuming df_range is your DataFrame
    #print(df_range_size)
    sample_size = round(df_range_size/2) if df_range_size < 250 else 250
    #print(sample_size)
    # Select 100 random cells within the specified range
    np.random.seed(0)  # for reproducibility
    random_cells = series.sample(sample_size).index.tolist()
    print(f"Writing Data to xlsx output_{timestr}.xlsx")
    excel_output =os.path.join(base_directory,f"output_{timestr}.xlsx")
    # Save DataFrame to an Excel file
    df.replace('', None, inplace=True)
    df = df.fillna('\\N')
    df.to_excel(excel_output, index=False)


    # Load the workbook and select the sheet
    book = load_workbook(excel_output)
    sheet = book.active  # adjust if your sheet has a different name
   

    # Create a fill object
    fill = PatternFill(fill_type="solid", fgColor="FFFF00")  # Yellow fill

    # Get column number from column label
    col_label_to_num = {col: num for num, col in enumerate(df.columns, start=1)}

    # Apply fill to selected cells
    for (row, col) in random_cells:
        col_num = col_label_to_num[col] # add 1 to adjust the offset of index between pandas and excel
        cell = sheet.cell(row=row+2, column=col_num)  # +2 to account for 0-indexing and header
        cell.fill = fill

    # Save the changes
    book.save(excel_output)

if __name__ == "__main__":
    # Assuming you have a DataFrame df
    timestr = '20230811204335'#20230803144127'
    input_lines = sys.stdin.readlines()
    if input_lines :
        for line in input_lines:
            input_data = line.strip()
            #print(f"Received input: {input_data}")
            timestr=input_data
    base_directory = os.path.join(os.getcwd(), "mappedoutput")
    input_file =os.path.join(base_directory,f"data_map_{timestr}.csv")
    df = pd.read_csv(input_file ,encoding='ISO-8859-1').fillna('')
    df = df.astype(str)
    #df =df.replace(r'^\s*$', np.nan, regex=True).replace([r'', r'\N'], np.nan)

    # List of column names to validate
    date_columns = ['Date','CapGainRecordDate','CapGainExDate','CapGainPayableDate']
    percentage_columns = ['ShortTermPercent','LongTermPercent','TotalCapGainsPercent', 'ShortTermPercentTop','LongTermPercentTop','TotalCapGainsPercentTop','ShortTermPercentLow','LongTermPercentLow','TotalCapGainsPercentLow']
    primary_columns = ['Provider']
    dollar_columns = ['ShortTerm','LongTerm','TotalCapGains','ShortTermTop','LongTermTop','TotalCapGainsTop','ShortTermLow','LongTermLow','TotalCapGainsLow']

    # Check that all dates are valid or null:
    for col in date_columns:
        df[f'is_{col}_valid'] = pd.to_datetime(df[col], format='%m/%d/%Y', errors='coerce').notna() | (df[col] == 'nan') | (df[col] == '')

    # Check that all percentages are valid or 0%:
    for col in percentage_columns:
        df[f'is_{col}_valid'] = df[col].str.endswith('%') & df[col].str.rstrip('%').apply(is_float) | (df[col] == '0%') | (df[col] == 'nan') | (df[col] == '')

    # Check if primary column is non-null or 0:
    for col in primary_columns:
        df[f'is_{col}_valid'] = df[col].notnull() | (df[col] == 0) 

    for col in dollar_columns:
        df[f'is_{col}_valid'] = df[col].str.startswith('$') & df[col].str.lstrip('$').apply(is_float) | (df[col] == '$0') | (df[col] == 'nan') | (df[col] == '')

    # Create a new column that is True if all validations passed and False otherwise
    df['is_row_valid'] = df.loc[:, df.columns.str.startswith('is_')].all(axis=1)

    # Separate good rows and bad rows into two dataframes
    good_df = df[df['is_row_valid']].copy()
    bad_df = df[~df['is_row_valid']].copy()

    # Remove validation columns from the good dataframe and the bad dataframe
    good_df.drop(columns=good_df.columns[good_df.columns.str.startswith('is_')], inplace=True)
    #bad_df.drop(columns=bad_df.columns[bad_df.columns.str.startswith('is_')], inplace=True)

    #print('number of good rows ',len(good_df))
    #print('number of bad rows ',len(bad_df))
    #print(bad_df)

    savetolocal(good_df,"good_data")
    savetolocal(bad_df,"bad_data")
    selectrandom(good_df)
    #print("RESULT: Success")
    #accuracy(good_df)

